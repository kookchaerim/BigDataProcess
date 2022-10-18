#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

list=[]
row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		list.append(sum_v)	
	row_id += 1

row_id = 1;
result = []

for i in list:
	rank = 1
	for j in list:
		if i < j:
			rank += 1
	result.append(rank)

row_id = 2;
tmp = len(result) / 10
for i in result:
	if i <= (tmp*3 / 2):
		ws.cell(row = row_id, column = 8).value = 'A+'
	elif i <= tmp*3:
		ws.cell(row = row_id, column = 8).value = 'A0'
	elif i <= tmp*3 + (tmp*7 - tmp*3) / 2:
		ws.cell(row = row_id, column = 8).value = 'B+'
	elif i <= tmp*7:
		ws.cell(row = row_id, column = 8).value = 'B0'
	elif i <= tmp*7 + (tmp*10 - tmp*7) / 2:
		ws.cell(row = row_id, column = 8).value = 'C+'
	else:
		ws.cell(row = row_id, column = 8).value = 'C0'
	row_id += 1
wb.save("student.xlsx")
